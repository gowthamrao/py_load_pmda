import openpyxl
from datetime import datetime

# Create a new workbook
wb = openpyxl.Workbook()

# Select the active worksheet
ws = wb.active

# Add some data
ws['A1'] = '承認番号'
ws['B1'] = '販売名'
ws['C1'] = '申請者'
ws['D1'] = '承認日'
ws['A2'] = '(302AMX00001000)'
ws['B2'] = 'テストメディカル'
ws['C2'] = 'テスト製薬株式会社'
ws['D2'] = datetime(2025, 9, 8)


# Save the workbook
wb.save("tests/fixtures/empty.xlsx")
